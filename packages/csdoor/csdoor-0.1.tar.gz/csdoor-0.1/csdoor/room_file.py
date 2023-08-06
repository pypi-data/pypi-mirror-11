from openpyxl import Workbook

class RoomFile:
    SHEET_TITLE = "Sheet1"
    FILE_NAME = "{file_path}/SUBgroup_{roomnumber}full.xlsx"

    def __init__(self, room_number, path):
        if path[-1] == '/':
            path = path[:-1]

        self.workbook, self.worksheet = self.create_sheet()
        self.filename = RoomFile.FILE_NAME.format(roomnumber=room_number,file_path=path)

    def create_sheet(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = RoomFile.SHEET_TITLE
        worksheet.append(["FirstName", "LastName", "ID_NUM", "SubGroup", "Action"])
        return workbook, worksheet
